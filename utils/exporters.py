import os
from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime

# ReportLab Imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# OpenPyXL Imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class Exporters:
    """Document generation utility for PDF and Excel exports."""

    @staticmethod
    def generate_pdf_report(employees: List[Dict[str, Any]], stats: Dict[str, Any]) -> BytesIO:
        """Generates a professional PDF report containing employee records and summary stats."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=colors.HexColor('#4f46e5'),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=25
        )
        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor('#1e1b4b'),
            spaceAfter=10
        )
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#1e293b')
        )
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white
        )

        # Header Info
        story.append(Paragraph("WorkSphere Enterprise Report", title_style))
        story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Scope: Company Workforce", subtitle_style))
        story.append(Spacer(1, 10))

        # KPI Cards Table
        story.append(Paragraph("Executive Summary", section_heading))
        summary_data = [
            [
                Paragraph("<b>Total Headcount:</b>", cell_style),
                Paragraph(str(stats.get('headcount', 0)), cell_style),
                Paragraph("<b>Average Salary:</b>", cell_style),
                Paragraph(f"${stats.get('avg_salary', 0.0):,.2f}", cell_style)
            ],
            [
                Paragraph("<b>Median Salary:</b>", cell_style),
                Paragraph(f"${stats.get('median_salary', 0.0):,.2f}", cell_style),
                Paragraph("<b>Attendance Rate:</b>", cell_style),
                Paragraph(f"{stats.get('attendance_stats', {}).get('mean_rate', 100.0):.1f}%", cell_style)
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[120, 120, 120, 120])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
            ('PADDING', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 25))

        # Employees Table
        story.append(Paragraph("Employee Details Directory", section_heading))
        
        table_headers = ["ID", "Name", "Email", "Department", "Designation", "Salary"]
        table_data = [[Paragraph(h, header_style) for h in table_headers]]

        for emp in employees:
            table_data.append([
                Paragraph(str(emp.get('emp_id', '')), cell_style),
                Paragraph(emp.get('name', ''), cell_style),
                Paragraph(emp.get('email', ''), cell_style),
                Paragraph(emp.get('department_name', 'Unassigned') or 'Unassigned', cell_style),
                Paragraph(emp.get('designation', '') or 'N/A', cell_style),
                Paragraph(f"${emp.get('salary', 0.0):,.2f}", cell_style),
            ])

        emp_table = Table(table_data, colWidths=[30, 95, 125, 100, 100, 70])
        emp_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(emp_table)

        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_excel_report(employees: List[Dict[str, Any]]) -> BytesIO:
        """Generates a styled Excel spreadsheet directory of employees."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Workforce Directory"

        # Design colors
        primary_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        font_bold = Font(name="Calibri", size=11, bold=True)
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        thin_border_side = Side(border_style="thin", color="E2E8F0")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Columns Headers
        headers = ["Employee ID", "Full Name", "Email Address", "Phone", "Department", "Designation", "Salary Status"]
        ws.append(headers)

        # Style Headers
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = primary_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        # Add Data
        for row_idx, emp in enumerate(employees, start=2):
            row_data = [
                emp.get('emp_id', ''),
                emp.get('name', ''),
                emp.get('email', ''),
                emp.get('phone', '') or 'N/A',
                emp.get('department_name', 'Unassigned') or 'Unassigned',
                emp.get('designation', '') or 'N/A',
                emp.get('salary', 0.0)
            ]
            ws.append(row_data)
            
            # Formatting cells
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                
                # Zebra Striping
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
                
                # Alignments and Number Formats
                if col_idx in [1, 4]:
                    cell.alignment = align_center
                elif col_idx == 7:
                    cell.alignment = align_right
                    cell.number_format = '$#,##0.00'
                else:
                    cell.alignment = align_left

        # Add Sum/Summary Row
        sum_row_idx = len(employees) + 2
        ws.cell(row=sum_row_idx, column=6, value="Total Payroll:").font = font_bold
        ws.cell(row=sum_row_idx, column=6).alignment = align_right
        
        salary_sum_cell = ws.cell(row=sum_row_idx, column=7, value=f"=SUM(G2:G{sum_row_idx-1})")
        salary_sum_cell.font = font_bold
        salary_sum_cell.number_format = '$#,##0.00'
        salary_sum_cell.border = Border(top=Side(style='thin', color="000000"), bottom=Side(style='double', color="000000"))

        # Adjust Columns widths dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format == '$#,##0.00' and isinstance(cell.value, (int, float)):
                    val = f"${cell.value:,.2f}"
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
